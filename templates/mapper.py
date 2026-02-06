"""
COREP template mapper - maps LLM output to Excel templates.
Creates DataFrames and exports to Excel format.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


class COREPTemplateMapper:
    """Maps analysis results to COREP Excel templates."""
    
    # Define template structure for C01.00 (Own Funds)
    C0100_STRUCTURE = {
        "rows": {
            "010": "Capital instruments and related share premium accounts",
            "020": "of which: ordinary shares",
            "030": "Retained earnings",
            "040": "Accumulated other comprehensive income",
            "050": "Other reserves",
            "060": "Funds for general banking risk",
            "070": "Minority interests",
            "080": "Transitional adjustments",
            "090": "Common Equity Tier 1 capital before deductions",
            "100": "Intangible assets",
            "110": "Deferred tax assets",
            "120": "Other deductions",
            "130": "Total deductions from CET1",
            "140": "Common Equity Tier 1 capital (CET1)",
            "150": "Additional Tier 1 instruments",
            "160": "Tier 1 capital (T1 = CET1 + AT1)",
            "170": "Tier 2 instruments",
            "180": "Total Own Funds (T1 + T2)"
        },
        "columns": {
            "010": "Amount",
            "020": "of which: CRR transitional rules",
            "030": "of which: Regulation (EU) No 575/2013"
        }
    }
    
    def __init__(self, template_code: str = "C01.00"):
        """
        Initialize the mapper.
        
        Args:
            template_code: COREP template code
        """
        self.template_code = template_code
        
    def create_dataframe_from_analysis(self, analysis: Dict) -> pd.DataFrame:
        """
        Create a DataFrame from LLM analysis result.
        
        Args:
            analysis: Analysis dictionary with 'fields' list
            
        Returns:
            Pandas DataFrame
        """
        # Initialize empty template structure
        rows = list(self.C0100_STRUCTURE["rows"].keys())
        columns = list(self.C0100_STRUCTURE["columns"].keys())
        
        # Create empty DataFrame
        df = pd.DataFrame(index=rows, columns=columns)
        df.index.name = "Row"
        
        # Fill with zeros
        df = df.fillna(0).infer_objects(copy=False)
        
        # Populate from analysis
        if "fields" in analysis:
            for field in analysis["fields"]:
                row_code = field.get("row")
                col_code = field.get("column")
                value = field.get("value", 0)
                
                if row_code in df.index and col_code in df.columns:
                    df.loc[row_code, col_code] = value
        
        return df
    
    def create_detailed_table(self, analysis: Dict) -> pd.DataFrame:
        """
        Create a detailed table with row labels and justifications.
        
        Args:
            analysis: Analysis dictionary
            
        Returns:
            DataFrame with labels and values
        """
        rows_data = []
        
        if "fields" in analysis:
            for field in analysis["fields"]:
                row_data = {
                    "Row": field.get("row"),
                    "Column": field.get("column"),
                    "Item": field.get("item_name", ""),
                    "Value": field.get("value", 0),
                    "Justification": field.get("justification", ""),
                    "Source": field.get("source", "")
                }
                rows_data.append(row_data)
        
        df = pd.DataFrame(rows_data)
        return df
    
    def export_to_excel(self, 
                       analysis: Dict,
                       output_path: str,
                       include_details: bool = True) -> str:
        """
        Export analysis to formatted Excel file.
        
        Args:
            analysis: Analysis dictionary
            output_path: Path for output Excel file
            include_details: Include detailed justifications sheet
            
        Returns:
            Path to created Excel file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Template values
            df_template = self.create_dataframe_from_analysis(analysis)
            df_template.to_excel(writer, sheet_name='C01.00_Template')
            
            # Sheet 2: Detailed breakdown
            if include_details:
                df_details = self.create_detailed_table(analysis)
                df_details.to_excel(writer, sheet_name='Details', index=False)
            
            # Sheet 3: Row descriptions
            row_descriptions = pd.DataFrame([
                {"Row Code": code, "Description": desc}
                for code, desc in self.C0100_STRUCTURE["rows"].items()
            ])
            row_descriptions.to_excel(writer, sheet_name='Row_Definitions', index=False)
        
        # Apply formatting
        self._format_excel(output_path)
        
        return str(output_path)
    
    def _format_excel(self, file_path: Path) -> None:
        """
        Apply formatting to Excel file.
        
        Args:
            file_path: Path to Excel file
        """
        wb = openpyxl.load_workbook(file_path)
        
        # Format template sheet
        if 'C01.00_Template' in wb.sheetnames:
            ws = wb['C01.00_Template']
            
            # Header formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Number formatting
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column > 1:  # Skip index column
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right")
        
        # Format details sheet
        if 'Details' in wb.sheetnames:
            ws = wb['Details']
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def create_summary_dataframe(self, analysis: Dict) -> pd.DataFrame:
        """
        Create a summary DataFrame suitable for display.
        
        Args:
            analysis: Analysis dictionary
            
        Returns:
            Summary DataFrame
        """
        df = self.create_detailed_table(analysis)
        
        # Format values with currency
        if 'Value' in df.columns:
            df['Value_Formatted'] = df['Value'].apply(lambda x: f"€{x:,.2f}" if pd.notnull(x) else "")
        
        return df


def map_to_template(analysis: Dict, output_path: Optional[str] = None) -> pd.DataFrame:
    """
    Convenience function to map analysis to template.
    
    Args:
        analysis: Analysis dictionary
        output_path: Optional path to save Excel file
        
    Returns:
        DataFrame
    """
    mapper = COREPTemplateMapper()
    
    if output_path:
        mapper.export_to_excel(analysis, output_path)
    
    return mapper.create_dataframe_from_analysis(analysis)


if __name__ == "__main__":
    # Test the mapper
    mock_analysis = {
        "template": "C01.00",
        "fields": [
            {
                "row": "010",
                "column": "010",
                "value": 10000000,
                "item_name": "Share capital",
                "justification": "Ordinary shares qualify as CET1",
                "source": "Own Funds (CRR), page 15"
            },
            {
                "row": "030",
                "column": "010",
                "value": 5000000,
                "item_name": "Retained earnings",
                "justification": "Retained earnings included in CET1",
                "source": "Own Funds (CRR), page 16"
            },
            {
                "row": "100",
                "column": "010",
                "value": -500000,
                "item_name": "Intangible assets",
                "justification": "Deducted from CET1",
                "source": "Own Funds (CRR), page 20"
            }
        ]
    }
    
    mapper = COREPTemplateMapper()
    
    # Create DataFrame
    df = mapper.create_dataframe_from_analysis(mock_analysis)
    print("Template DataFrame:")
    print(df)
    
    # Create detailed table
    df_details = mapper.create_detailed_table(mock_analysis)
    print("\nDetailed Table:")
    print(df_details)
    
    # Export to Excel
    output_file = "test_output.xlsx"
    path = mapper.export_to_excel(mock_analysis, output_file)
    print(f"\nExcel file created: {path}")
